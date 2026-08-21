"""Follow-up workbook generation.

The deck is for the room; this workbook is the working list. It has two
sheets, both answering "what is wrong with the board" rather than "how did
the sprint go":

``Unestimated``
    Items touched during the sprint that carry no estimate. These are the
    reason a points figure can understate delivery — an item that shipped
    without an estimate contributes nothing to completed points.

``Off Sprint``
    Issues that saw activity in the sprint window but were never assigned to
    the iteration, plus issues that are not on the project board at all. Work
    happening entirely outside the process is invisible to every report.

Both sheets are plain data — no formulas — so they can be sorted, filtered,
and worked down directly.

Example:
    >>> from sprint_report.models import ProjectItem
    >>> import tempfile, pathlib
    >>> with tempfile.TemporaryDirectory() as tmp:
    ...     path = build_workbook(
    ...         pathlib.Path(tmp) / "followups.xlsx",
    ...         iteration="Sprint 1",
    ...         unestimated=[ProjectItem(item_id="1", title="t")],
    ...         off_sprint=[],
    ...     )
    ...     path.exists()
    True
"""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import ProjectItem

__all__ = ["build_workbook", "OffSprintIssue"]

#: Body font used throughout, per the house spreadsheet convention.
BODY_FONT = "Arial"

#: Header fill, matching the deck's ink colour.
HEADER_FILL = PatternFill("solid", fgColor="1F2933")

#: Hyperlink colour, matching the deck's accent.
LINK_COLOUR = "0F766E"


class OffSprintIssue:
    """An issue that saw activity without being tracked in the sprint.

    Attributes:
        number: Issue number.
        title: Issue title.
        url: Issue URL.
        state: ``open`` or ``closed``.
        updated: ISO date the issue was last updated.
        assignees: Comma-separated assignee logins.
        on_board: Whether the issue is on the project board but unsprinted
            (``True``) or absent from the board entirely (``False``).

    Example:
        >>> OffSprintIssue(21, "Routes", "https://x/21", "open",
        ...                "2026-08-14", "", True).reason
        'On board, no sprint'
    """

    def __init__(
        self,
        number: int | str,
        title: str,
        url: str,
        state: str,
        updated: str,
        assignees: str,
        on_board: bool,
    ) -> None:
        """Initialise the record. See class docstring for attribute meanings."""
        self.number = number
        self.title = title
        self.url = url
        self.state = state
        self.updated = updated
        self.assignees = assignees
        self.on_board = on_board

    @property
    def reason(self) -> str:
        """Why this issue appears on the sheet.

        Returns:
            A short phrase distinguishing the two cases.

        Example:
            >>> OffSprintIssue(1, "t", "", "open", "", "", False).reason
            'Not on the board'
        """
        return "On board, no sprint" if self.on_board else "Not on the board"


def _write_header(sheet: Worksheet, headers: Sequence[str]) -> None:
    """Write and style a header row.

    Args:
        sheet: Target worksheet.
        headers: Column titles, left to right.
    """
    for column, label in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=label)
        cell.font = Font(name=BODY_FONT, size=11, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"


def _style_row(sheet: Worksheet, row: int, columns: int) -> None:
    """Apply the body font across a row.

    Args:
        sheet: Target worksheet.
        row: 1-indexed row number.
        columns: How many columns to style.
    """
    for column in range(1, columns + 1):
        cell = sheet.cell(row=row, column=column)
        if column != 1:
            cell.font = Font(name=BODY_FONT, size=11)
        cell.alignment = Alignment(vertical="top", wrap_text=column == 2)


def _link(sheet: Worksheet, row: int, column: int, text: str, url: str) -> None:
    """Write a cell as a hyperlink when a URL is available.

    Args:
        sheet: Target worksheet.
        row: 1-indexed row.
        column: 1-indexed column.
        text: Display text.
        url: Target URL; a plain cell is written when empty.
    """
    cell = sheet.cell(row=row, column=column, value=text)
    if url:
        cell.hyperlink = url
        cell.font = Font(name=BODY_FONT, size=11, color=LINK_COLOUR, underline="single")
    else:
        cell.font = Font(name=BODY_FONT, size=11)


def _autosize(sheet: Worksheet, widths: Sequence[int]) -> None:
    """Set fixed column widths.

    Args:
        sheet: Target worksheet.
        widths: Width per column, left to right.
    """
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _unestimated_sheet(
    sheet: Worksheet, iteration: str, items: Sequence[ProjectItem]
) -> None:
    """Populate the unestimated-items sheet.

    Args:
        sheet: Target worksheet.
        iteration: Iteration title, used in the note row.
        items: Items in the sprint carrying no estimate.
    """
    sheet.title = "Unestimated"
    _write_header(sheet, ("Issue", "Title", "Status", "Assignees", "Repository"))
    _autosize(sheet, (10, 70, 16, 24, 26))

    for row, item in enumerate(items, start=2):
        _link(sheet, row, 1, f"#{item.item_id}", item.url)
        sheet.cell(row=row, column=2, value=item.title)
        sheet.cell(row=row, column=3, value=item.status or "—")
        sheet.cell(row=row, column=4, value="")
        sheet.cell(row=row, column=5, value=item.repository)
        _style_row(sheet, row, 5)

    note_row = len(items) + 3
    note = sheet.cell(
        row=note_row,
        column=1,
        value=(
            f"Items in {iteration} with no value in the estimate field. Each "
            "counts as zero points, so completed-points figures understate "
            "delivery until these are filled in."
        ),
    )
    note.font = Font(name=BODY_FONT, size=10, italic=True, color="616E7C")


def _off_sprint_sheet(
    sheet: Worksheet,
    iteration: str,
    issues: Sequence[OffSprintIssue],
) -> None:
    """Populate the off-sprint sheet.

    Args:
        sheet: Target worksheet.
        iteration: Iteration title, used in the note row.
        issues: Issues worked on outside the iteration.
    """
    sheet.title = "Off Sprint"
    _write_header(
        sheet,
        ("Issue", "Title", "State", "Last updated", "Assignees", "Why it's here"),
    )
    _autosize(sheet, (10, 62, 12, 16, 24, 22))

    for row, issue in enumerate(issues, start=2):
        _link(sheet, row, 1, f"#{issue.number}", issue.url)
        sheet.cell(row=row, column=2, value=issue.title)
        sheet.cell(row=row, column=3, value=issue.state)
        sheet.cell(row=row, column=4, value=issue.updated)
        sheet.cell(row=row, column=5, value=issue.assignees)
        sheet.cell(row=row, column=6, value=issue.reason)
        _style_row(sheet, row, 6)

    note_row = len(issues) + 3
    note = sheet.cell(
        row=note_row,
        column=1,
        value=(
            f"Issues updated during {iteration} that were never assigned to "
            "the iteration, or are not on the project board at all. This work "
            "appears in no sprint report."
        ),
    )
    note.font = Font(name=BODY_FONT, size=10, italic=True, color="616E7C")


def build_workbook(
    output_path: Path,
    iteration: str,
    unestimated: Sequence[ProjectItem],
    off_sprint: Sequence[OffSprintIssue],
) -> Path:
    """Write the two-sheet follow-up workbook.

    Args:
        output_path: Destination ``.xlsx`` path. Parent directories are
            created as needed.
        iteration: Iteration title, used in the explanatory note rows.
        unestimated: Sprint items carrying no estimate.
        off_sprint: Issues worked on outside the iteration.

    Returns:
        The path written.

    Raises:
        OSError: If the file cannot be written.

    Example:
        >>> import tempfile, pathlib
        >>> with tempfile.TemporaryDirectory() as tmp:
        ...     p = build_workbook(
        ...         pathlib.Path(tmp) / "f.xlsx", "Sprint 1", [], [])
        ...     p.name
        'f.xlsx'
    """
    workbook = Workbook()
    _unestimated_sheet(workbook.active, iteration, unestimated)
    _off_sprint_sheet(workbook.create_sheet(), iteration, off_sprint)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(output_path))
    return output_path

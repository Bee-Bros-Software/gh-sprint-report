"""Tests for :mod:`sprint_report.workbook` and off-sprint discovery."""

from __future__ import annotations

import argparse
import subprocess
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from sprint_report.models import ProjectItem, SprintMetrics
from sprint_report.workbook import OffSprintIssue, build_workbook


class TestOffSprintIssue:
    """The record type's derived reason."""

    def test_on_board_reason(self):
        """A board item with no iteration says so."""
        record = OffSprintIssue(21, "Routes", "", "open", "", "", True)
        assert record.reason == "On board, no sprint"

    def test_off_board_reason(self):
        """An issue absent from the board says so."""
        record = OffSprintIssue(99, "x", "", "open", "", "", False)
        assert record.reason == "Not on the board"


class TestWorkbook:
    """Structure and content of the generated file."""

    def _build(self, tmp_path: Path, unestimated=(), off_sprint=()) -> Path:
        """Generate a workbook.

        Args:
            tmp_path: pytest temporary directory.
            unestimated: Items for the first sheet.
            off_sprint: Records for the second sheet.

        Returns:
            The generated path.
        """
        return build_workbook(
            tmp_path / "followups.xlsx", "Sprint 1", unestimated, off_sprint
        )

    def test_has_both_sheets_named_exactly(self, tmp_path: Path):
        """Sheet names are fixed so downstream filters can rely on them."""
        book = load_workbook(self._build(tmp_path))
        assert book.sheetnames == ["Unestimated", "Off Sprint"]

    def test_unestimated_rows(self, tmp_path: Path):
        """Each unestimated item becomes one row under the header."""
        items = [
            ProjectItem(
                item_id="48",
                title="PHI-access audit log",
                url="https://github.com/acme/r/issues/48",
                status="Done",
                repository="acme/r",
            )
        ]
        sheet = load_workbook(self._build(tmp_path, unestimated=items))["Unestimated"]
        assert sheet["A1"].value == "Issue"
        assert sheet["A2"].value == "#48"
        assert sheet["B2"].value == "PHI-access audit log"
        assert sheet["C2"].value == "Done"

    def test_issue_cell_is_hyperlinked(self, tmp_path: Path):
        """The issue number links back to GitHub."""
        items = [
            ProjectItem(
                item_id="48", title="t", url="https://github.com/acme/r/issues/48"
            )
        ]
        sheet = load_workbook(self._build(tmp_path, unestimated=items))["Unestimated"]
        assert sheet["A2"].hyperlink.target == "https://github.com/acme/r/issues/48"

    def test_item_without_url_is_plain_text(self, tmp_path: Path):
        """A draft item with no URL still renders, without a broken link."""
        items = [ProjectItem(item_id="1", title="draft")]
        sheet = load_workbook(self._build(tmp_path, unestimated=items))["Unestimated"]
        assert sheet["A2"].value == "#1"
        assert sheet["A2"].hyperlink is None

    def test_off_sprint_rows(self, tmp_path: Path):
        """Off-sprint records carry state, date, assignees, and a reason."""
        records = [
            OffSprintIssue(
                21, "Routes", "https://x/21", "open", "2026-08-14", "tjirousek", True
            )
        ]
        sheet = load_workbook(self._build(tmp_path, off_sprint=records))["Off Sprint"]
        assert sheet["A2"].value == "#21"
        assert sheet["C2"].value == "open"
        assert sheet["D2"].value == "2026-08-14"
        assert sheet["E2"].value == "tjirousek"
        assert sheet["F2"].value == "On board, no sprint"

    def test_explanatory_note_present(self, tmp_path: Path):
        """Each sheet explains what it is, below the data."""
        book = load_workbook(self._build(tmp_path))
        assert "understate delivery" in book["Unestimated"]["A3"].value
        assert "appears in no sprint report" in book["Off Sprint"]["A3"].value

    def test_header_is_frozen(self, tmp_path: Path):
        """Long lists stay readable when scrolled."""
        book = load_workbook(self._build(tmp_path))
        assert book["Unestimated"].freeze_panes == "A2"

    def test_uses_a_professional_font(self, tmp_path: Path):
        """Arial throughout, per the house convention."""
        items = [ProjectItem(item_id="1", title="t")]
        sheet = load_workbook(self._build(tmp_path, unestimated=items))["Unestimated"]
        assert sheet["B2"].font.name == "Arial"
        assert sheet["A1"].font.bold

    def test_empty_workbook_still_valid(self, tmp_path: Path):
        """Nothing to follow up on produces headers and notes, not an error."""
        book = load_workbook(self._build(tmp_path))
        assert book["Unestimated"]["A1"].value == "Issue"

    def test_creates_parent_directory(self, tmp_path: Path):
        """A nested output path is created rather than failing."""
        target = tmp_path / "deep" / "nested" / "f.xlsx"
        assert build_workbook(target, "S1", [], []).exists()


class TestOffSprintDiscovery:
    """Assembly of the off-sprint population in the CLI."""

    def _args(self, **overrides) -> argparse.Namespace:
        """Build a namespace with discovery defaults.

        Args:
            **overrides: Fields to override.

        Returns:
            An argparse-like namespace.
        """
        base = {"issues_repo": None}
        base.update(overrides)
        return argparse.Namespace(**base)

    def test_includes_worked_board_items_without_a_sprint(self):
        """A board item in progress but unsprinted is off-sprint work."""
        from sprint_report.cli import _off_sprint_issues

        items = [ProjectItem(item_id="46", title="t", status="Done")]
        records = _off_sprint_issues(
            self._args(), items, SprintMetrics("S1")
        )
        assert [r.number for r in records] == ["46"]
        assert records[0].on_board is True

    def test_excludes_untouched_backlog(self):
        """A Todo item nobody started is backlog, not off-sprint work."""
        from sprint_report.cli import _off_sprint_issues

        items = [ProjectItem(item_id="21", title="t", status="Todo")]
        assert _off_sprint_issues(self._args(), items, SprintMetrics("S1")) == []

    def test_excludes_items_already_in_a_sprint(self):
        """Sprinted work is not off-sprint by definition."""
        from sprint_report.cli import _off_sprint_issues

        items = [
            ProjectItem(item_id="48", title="t", status="Done", iteration="S1")
        ]
        assert _off_sprint_issues(self._args(), items, SprintMetrics("S1")) == []

    def test_repo_scan_skipped_without_a_start_date(self):
        """No date window means no repository query to make."""
        from sprint_report.cli import _off_sprint_issues

        items = [ProjectItem(item_id="1", title="t", repository="acme/r")]
        assert _off_sprint_issues(self._args(), items, SprintMetrics("S1")) == []

    def test_repo_scan_adds_off_board_issues(self, monkeypatch):
        """An issue absent from the board is surfaced with its metadata."""
        from sprint_report import cli

        monkeypatch.setattr(
            cli,
            "fetch_issues",
            lambda repo, since, **k: [
                {
                    "number": 99,
                    "title": "Hotfix nobody tracked",
                    "url": "https://x/99",
                    "state": "CLOSED",
                    "updatedAt": "2026-08-15T09:00:00Z",
                    "assignees": [{"login": "N0rvil"}],
                }
            ],
        )
        items = [
            ProjectItem(
                item_id="48",
                title="t",
                iteration="S1",
                repository="acme/r",
            )
        ]
        metrics = SprintMetrics("S1", date(2026, 8, 10), date(2026, 8, 23))
        records = cli._off_sprint_issues(self._args(), items, metrics)
        assert len(records) == 1
        assert records[0].number == "99"
        assert records[0].state == "closed"
        assert records[0].updated == "2026-08-15"
        assert records[0].assignees == "N0rvil"
        assert records[0].on_board is False

    def test_repo_scan_ignores_issues_already_on_the_board(self, monkeypatch):
        """An issue on the board is not also reported as off-board."""
        from sprint_report import cli

        monkeypatch.setattr(
            cli,
            "fetch_issues",
            lambda repo, since, **k: [
                {"number": 48, "title": "t", "url": "", "state": "open"}
            ],
        )
        items = [
            ProjectItem(item_id="48", title="t", iteration="S1", repository="acme/r")
        ]
        metrics = SprintMetrics("S1", date(2026, 8, 10), date(2026, 8, 23))
        assert cli._off_sprint_issues(self._args(), items, metrics) == []

    def test_gh_failure_degrades_to_board_data(self, monkeypatch, capsys):
        """A failed repo scan warns but does not fail the whole report."""
        from sprint_report import cli
        from sprint_report.gh_source import GhError

        def _boom(repo, since, **k):
            raise GhError("no such repo")

        monkeypatch.setattr(cli, "fetch_issues", _boom)
        items = [
            ProjectItem(item_id="1", title="t", iteration="S1", repository="acme/r")
        ]
        metrics = SprintMetrics("S1", date(2026, 8, 10), date(2026, 8, 23))
        assert cli._off_sprint_issues(self._args(), items, metrics) == []
        assert "Skipping acme/r" in capsys.readouterr().err

    def test_repository_urls_are_normalised(self, monkeypatch):
        """Board rows carry full URLs; gh needs owner/name."""
        from sprint_report import cli

        seen: list[str] = []
        monkeypatch.setattr(
            cli,
            "fetch_issues",
            lambda repo, since, **k: seen.append(repo) or [],
        )
        items = [
            ProjectItem(
                item_id="1",
                title="t",
                iteration="S1",
                repository="https://github.com/acme/widgets",
            )
        ]
        metrics = SprintMetrics("S1", date(2026, 8, 10), date(2026, 8, 23))
        cli._off_sprint_issues(self._args(), items, metrics)
        assert seen == ["acme/widgets"]


class TestFetchIssues:
    """The gh issue list wrapper."""

    def test_parses_issue_json(self, monkeypatch):
        """A zero exit is parsed into raw issue dictionaries."""
        import json

        from sprint_report.gh_source import fetch_issues

        monkeypatch.setattr("shutil.which", lambda _: "/usr/bin/gh")
        monkeypatch.setattr(
            subprocess,
            "run",
            lambda *a, **k: subprocess.CompletedProcess(
                a[0], 0, json.dumps([{"number": 1, "title": "t"}]), ""
            ),
        )
        assert fetch_issues("acme/r", "2026-08-10")[0]["number"] == 1

    def test_search_window_is_passed(self, monkeypatch):
        """The since date becomes a gh search qualifier."""
        from sprint_report.gh_source import fetch_issues

        monkeypatch.setattr("shutil.which", lambda _: "/usr/bin/gh")
        seen: list[list[str]] = []
        monkeypatch.setattr(
            subprocess,
            "run",
            lambda cmd, **k: seen.append(cmd)
            or subprocess.CompletedProcess(cmd, 0, "[]", ""),
        )
        fetch_issues("acme/r", "2026-08-10")
        assert "updated:>=2026-08-10" in seen[0]

    def test_failure_raises(self, monkeypatch):
        """A non-zero exit is reported, not swallowed."""
        from sprint_report.gh_source import GhError, fetch_issues

        monkeypatch.setattr("shutil.which", lambda _: "/usr/bin/gh")
        monkeypatch.setattr(
            subprocess,
            "run",
            lambda *a, **k: subprocess.CompletedProcess(a[0], 1, "", "boom"),
        )
        with pytest.raises(GhError, match="boom"):
            fetch_issues("acme/r", "2026-08-10")

    def test_invalid_json_raises(self, monkeypatch):
        """Malformed output fails loudly."""
        from sprint_report.gh_source import GhError, fetch_issues

        monkeypatch.setattr("shutil.which", lambda _: "/usr/bin/gh")
        monkeypatch.setattr(
            subprocess,
            "run",
            lambda *a, **k: subprocess.CompletedProcess(a[0], 0, "{bad", ""),
        )
        with pytest.raises(GhError, match="invalid JSON"):
            fetch_issues("acme/r", "2026-08-10")
